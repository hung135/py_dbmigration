import yaml
import logging
 
import sys
import pandas as pd
import numpy as np
import py_dbutils.rdbms.postgres as db_utils
import py_dbmigration.data_file_mgnt as data_file_mgnt
import py_dbmigration.migrate_utils as migrate_utils
from py_dbmigration.data_file_mgnt.state import *
import py_dbmigration.db_logging as db_logging
import py_dbmigration.db_table as db_table
from openpyxl import load_workbook

import pprint
import os, logging as log
runtime_pid=os.getpid()
logging = log.getLogger(f'\tPID: {runtime_pid} - {os.path.basename(__file__)}\t')
logging.setLevel(log.DEBUG)

HEADER_location = 0
UNITS_location = 1
COLUMN_location = 2
column_name_row_offset = 1

# parse the table of content sheet to return a list of start and end row location for each dataset


def process_tbl_content(wb):
    # wb = load_workbook(filename=file)
    data_dict = {}
    curr_category = 'None'
    prev_category = 'None'
    locations = {}
    # for sheet in wb.sheetnames:
    # Taqble of Content must be the first SHEEET!!!!!
    ws = wb.worksheets[0]
    # ws = wb[sheet]
    for row in ws.rows:

        for cell in row:
            val = cell.value

            if val is not None:

                link = cell.hyperlink
                if link is not None:

                    url = cell.hyperlink.location
                    # pprint.pprint(cell.hyperlink.location)
                    data_dict[curr_category].append(url)
                else:
                    prev_category = curr_category
                    curr_category = val
                    data_dict[curr_category] = []
    tbl_content = {}
    for data_loc in data_dict.items():

        name, value = data_loc
        if len(value) > 0:
            tbl_content[name] = value

    return tbl_content


def derive_start_end_rows(t):
    dict_items = []
    for item in t.items():
        sheetname, url_list = item
        start_row = None
        prev_loc = None
        for location in url_list:

            end_row = int(location.split('!')[1].split('$')[2]) - 1

            # print(cell_loc)
            # sheet = wb[sheetname]
            # process_worksheet(sheet, cell_loc)
            # print(start_row, end_row)
            if start_row is not None:
                dict_items.append({'sheetname': sheetname, 'start_row': start_row, 'end_row': end_row - 1})
            start_row = end_row
        dict_items.append({'sheetname': sheetname, 'start_row': start_row, 'end_row': None})

    return dict_items


def process_worksheet(sheet, start_row, end_row, has_units=True):
    # logging.debug('Processing Sheet: {} - Start Row: {} End Row: {}'.format(sheet, start_row, end_row))
    i = 0
    # for sheet in wb.sheetnames:
    # Table of Content must be the first SHEEET!!!!!
    if end_row is None:
        last_row = sheet.max_row
    else:
        last_row = end_row
    data_dict = {}
    data_dict['data'] = []
    data_dict['data_list'] = []
    data_dict['columns'] = None
    data_dict['unit'] = 'None'
    keep_nulls = False
    last_data_cell = 200
    for id, row in enumerate(sheet):
        # operating in the range of rows provided by the table of content

        if start_row <= id < last_row:
            line = []
            all_null = True

            first_column_w_data = 0
            for cell_number, cell in enumerate(row):

                if cell.value is None:
                    if keep_nulls:
                        line.append('"{}"'.format(str('NULL').encode('utf-8')))
                    pass
                elif cell.data_type == 's':
                    line.append('"{}"'.format(cell.value.encode('utf-8').replace('"', '""')))
                    all_null = False
                    if first_column_w_data == 0:
                        first_column_w_data = cell_number
                else:
                    line.append('"{}"'.format(str(cell.value).encode('utf-8')))
                    all_null = False
                    if first_column_w_data == 0:
                        first_column_w_data = cell_number

            # strip out leading and trailing blank cells to line up with columns
            if data_dict['columns'] is not None:
                last_data_cell = first_column_w_data + len(data_dict['columns'].split(','))
                new_line = []
                for id, cell in enumerate(line):
                    if first_column_w_data <= id < last_data_cell:
                        new_line.append(cell)

            line_str = ''
            if not all_null:
                line_str = ','.join(line)
            # skip lines that have nothing
            if len(line_str) > 1:
                if i == HEADER_location:
                    data_dict['header'] = line_str
                elif (i == UNITS_location and has_units):
                    data_dict['unit'] = line_str
                elif (i == COLUMN_location and has_units) or ((i == COLUMN_location - 1) and not has_units):
                    data_dict['columns'] = 'dimension",' + line_str
                    keep_nulls = True
                    data_dict['data_start_cell'] = first_column_w_data - 1

                else:

                    data_dict['data'].append(line_str)
                    data_dict['data_list'].append(new_line)

                i += 1
    return data_dict


# def process(db, file, file_id, dbschema):
def process(db, foi, df):
    continue_processing = True
    total_rows = 0
    target_dbschema = foi.schema_name
    table_name = foi.table_name
    file_id = df.meta_source_file_id
    file = os.path.join(df.source_file_path, df.curr_src_working_file)
    # file = '/home/dtdata/source_data/mkts_derived_data_warehouse/experian_mir/text/Experian-Oliver Wyman MIR - Student Loan DataPack.xlsx'
    with open(file, 'rb') as f:

        wb = load_workbook(filename=f)
        t = process_tbl_content(wb)
        t2 = derive_start_end_rows(t)
        # db = db_utils.DB(dbschema='autocount', dbtype='POSTGRES')
        sqlalchemy_conn, meta = db.connect_SqlAlchemy()
        # pprint.pprint(t2)
        prev_sheet = None
        for sheet in t2:

            sheetname = sheet['sheetname']
            if prev_sheet != sheetname:
                logging.info("\t\tProcessing Sheet: {}".format(sheetname))
                prev_sheet = sheetname
            has_units = True
            if sheetname == 'Snapshots':
                has_units = False

            d = process_worksheet(wb[sheetname], sheet['start_row'], sheet['end_row'], has_units)

            df = pd.DataFrame(d['data_list'], columns=d['columns'].split(','))
            df['sheetname'] = sheetname
            df['figure'] = d['header']
            df['units'] = d['unit']
            df['file_id'] = str(file_id)
            rm_quote = lambda x: x.replace('"', '')
            df = df.rename(columns=rm_quote)
            df = df.applymap(rm_quote)

            # dd = df.set_index(['sheetname', 'figure', 'units', '"Dimension"'])
            ee = pd.melt(df, id_vars=['file_id', 'sheetname', 'figure', 'units',
                                      'dimension'], var_name='measure', value_name='stat_value')

            ee = ee[ee.stat_value != 'NULL']
            # ee = ee.replace(to_replace=r'^-$', value=np.nan, regex=True)
            ee.to_sql(table_name, sqlalchemy_conn, schema=target_dbschema, if_exists='append', index=False)
             
            total_rows += len(ee)
        file_name=os.path.basename(__file__)
        t = db_table.db_table_func.RecordKeeper(db, db_table.db_table_def.MetaSourceFiles,file_name)
        row = t.get_record(db_table.db_table_def.MetaSourceFiles.id == file_id)
        row.total_rows = total_rows
        row.database_table = target_dbschema + '.' + table_name
        t.session.commit()
        t.session.close()

        logging.info("\t\tTotal Rows For this Excel File: {}".format(total_rows))
        return continue_processing
